"""Big Ambitions transactions.csv izleme + Google Drive senkronizasyon scripti.

Yeni davranış:
- Drive üzerinde `big ambitions` adlı kök klasörü garanti edilir.
- Gün değerine göre 60 günlük dönem klasörleri (1-60, 61-120, ...) oluşturulur.
- transactions.csv değiştiğinde ilgili dönem klasörüne transactionsgun_<gun>.csv yüklenir.
- Her dönem klasöründe `main.xlsx` üretilir (Excel SUMIF formülleri ile tip bazlı toplam).
- Kök klasörde `main_total.xlsx` üretilir (tüm dönemlerin toplamı, yine Excel formülleri).
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import tempfile
import time
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Optional

import psutil
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from openpyxl import Workbook
from watchdog.events import FileSystemEvent, FileSystemEventHandler
from watchdog.observers import Observer

SCOPES = ["https://www.googleapis.com/auth/drive"]
FOLDER_MIME = "application/vnd.google-apps.folder"
ROOT_FOLDER_NAME = "big ambitions"


@dataclass
class Config:
    savegames_root: Path
    service_account_file: Path
    process_names: tuple[str, ...]
    poll_seconds: int = 15
    file_settle_seconds: int = 10


class DriveClient:
    """Drive üzerinde klasör/dosya yönetimi ve upload/download yardımcıları."""

    def __init__(self, service_account_file: Path) -> None:
        credentials = service_account.Credentials.from_service_account_file(
            str(service_account_file), scopes=SCOPES
        )
        self.service = build("drive", "v3", credentials=credentials, cache_discovery=False)

    @staticmethod
    def _escape(name: str) -> str:
        return name.replace("'", "\\'")

    def _find_folder(self, name: str, parent_id: Optional[str]) -> Optional[str]:
        query = [
            f"name = '{self._escape(name)}'",
            f"mimeType = '{FOLDER_MIME}'",
            "trashed = false",
        ]
        if parent_id:
            query.append(f"'{parent_id}' in parents")
        response = (
            self.service.files()
            .list(q=" and ".join(query), fields="files(id,name)", pageSize=1)
            .execute()
        )
        files = response.get("files", [])
        return files[0]["id"] if files else None

    def ensure_folder(self, name: str, parent_id: Optional[str]) -> str:
        existing = self._find_folder(name, parent_id)
        if existing:
            return existing
        body = {"name": name, "mimeType": FOLDER_MIME}
        if parent_id:
            body["parents"] = [parent_id]
        created = self.service.files().create(body=body, fields="id").execute()
        return created["id"]

    def upload_or_update_file(
        self,
        local_file: Path,
        drive_name: str,
        parent_id: str,
        mimetype: str,
    ) -> str:
        file_id = self._find_file_id(drive_name, parent_id)
        media = MediaFileUpload(str(local_file), mimetype=mimetype, resumable=False)
        if file_id:
            self.service.files().update(fileId=file_id, media_body=media).execute()
            return f"Updated: {drive_name}"

        body = {"name": drive_name, "parents": [parent_id]}
        self.service.files().create(body=body, media_body=media, fields="id").execute()
        return f"Created: {drive_name}"

    def _find_file_id(self, name: str, parent_id: str) -> Optional[str]:
        query = [
            f"name = '{self._escape(name)}'",
            "trashed = false",
            f"'{parent_id}' in parents",
        ]
        response = (
            self.service.files()
            .list(q=" and ".join(query), fields="files(id,name)", pageSize=1)
            .execute()
        )
        files = response.get("files", [])
        return files[0]["id"] if files else None

    def list_files(self, parent_id: str, mime_exclude_folder: bool = False) -> list[dict]:
        query = ["trashed = false", f"'{parent_id}' in parents"]
        if mime_exclude_folder:
            query.append(f"mimeType != '{FOLDER_MIME}'")
        response = (
            self.service.files()
            .list(q=" and ".join(query), fields="files(id,name,mimeType)", pageSize=1000)
            .execute()
        )
        return response.get("files", [])

    def list_folders(self, parent_id: str) -> list[dict]:
        query = [
            "trashed = false",
            f"'{parent_id}' in parents",
            f"mimeType = '{FOLDER_MIME}'",
        ]
        response = (
            self.service.files()
            .list(q=" and ".join(query), fields="files(id,name)", pageSize=1000)
            .execute()
        )
        return response.get("files", [])

    def download_file(self, file_id: str, target_path: Path) -> None:
        request = self.service.files().get_media(fileId=file_id)
        with target_path.open("wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()


class ExcelAggregator:
    """Toplamları Excel formülleriyle oluşturan rapor üretici."""

    TYPE_COL_IDX = 2  # C sütunu
    VALUE_COL_IDX = 3  # D sütunu

    @classmethod
    def create_main_workbook(cls, csv_files: list[Path], output_path: Path, sheet_name: str) -> None:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Transactions"
        ws_data.append(["A", "B", "Type(C)", "Value(D)"])

        types: set[str] = set()
        for csv_path in csv_files:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                first = next(reader, None)
                if first is None:
                    continue

                # Başlık satırı olabilir.
                rows = reader if not (len(first) > 3 and cls._is_number(first[1])) else [first, *reader]
                for row in rows:
                    if len(row) <= cls.VALUE_COL_IDX:
                        continue
                    ws_data.append([
                        row[0] if len(row) > 0 else "",
                        row[1] if len(row) > 1 else "",
                        row[2] if len(row) > 2 else "",
                        cls._to_float_or_raw(row[3]),
                    ])
                    t = (row[2] if len(row) > 2 else "").strip()
                    if t:
                        types.add(t)

        ws_main = wb.create_sheet(sheet_name)
        ws_main.append(["Type", "Total"])

        sorted_types = sorted(types)
        for idx, expense_type in enumerate(sorted_types, start=2):
            ws_main.cell(row=idx, column=1, value=expense_type)
            ws_main.cell(
                row=idx,
                column=2,
                value=f"=SUMIF(Transactions!$C:$C,A{idx},Transactions!$D:$D)",
            )

        total_row = len(sorted_types) + 2
        ws_main.cell(row=total_row, column=1, value="GRAND_TOTAL")
        ws_main.cell(row=total_row, column=2, value=f"=SUM(B2:B{max(total_row - 1, 2)})")

        wb.save(output_path)

    @staticmethod
    def _is_number(value: str) -> bool:
        try:
            float(value.strip())
            return True
        except Exception:
            return False

    @staticmethod
    def _to_float_or_raw(value: str):
        raw = value.strip()
        try:
            return float(raw)
        except Exception:
            return raw


class TransactionsHandler(FileSystemEventHandler):
    def __init__(self, drive: DriveClient, settle_seconds: int = 10) -> None:
        self.drive = drive
        self.settle_seconds = settle_seconds
        self._last_uploaded_mtime: Optional[float] = None

    def on_modified(self, event: FileSystemEvent) -> None:
        if event.is_directory:
            return
        changed_file = Path(event.src_path)
        if changed_file.name.lower() != "transactions.csv":
            return

        try:
            print(f"[WATCHDOG] Değişim: {changed_file}")
            time.sleep(self.settle_seconds)
            file_mtime = changed_file.stat().st_mtime
            if self._last_uploaded_mtime == file_mtime:
                return

            day_value = extract_day_from_csv(changed_file)
            if day_value is None:
                print("[WARN] Gün bilgisi okunamadı.")
                return

            self._sync_transaction(changed_file, day_value)
            self._last_uploaded_mtime = file_mtime
        except Exception as exc:
            print(f"[ERROR] Event işleme hatası: {exc}")

    def _sync_transaction(self, source_csv: Path, day_value: int) -> None:
        root_id = self.drive.ensure_folder(ROOT_FOLDER_NAME, None)
        period_label = day_to_period_label(day_value)
        period_folder_id = self.drive.ensure_folder(period_label, root_id)

        drive_csv_name = f"transactionsgun_{day_value}.csv"
        tmp_csv = Path(tempfile.gettempdir()) / drive_csv_name
        shutil.copy2(source_csv, tmp_csv)
        try:
            print("[DRIVE]", self.drive.upload_or_update_file(tmp_csv, drive_csv_name, period_folder_id, "text/csv"))
        finally:
            tmp_csv.unlink(missing_ok=True)

        self._rebuild_period_main(period_folder_id)
        self._rebuild_global_main(root_id)

    def _rebuild_period_main(self, period_folder_id: str) -> None:
        csv_files = download_csv_files_from_folder(self.drive, period_folder_id)
        output = Path(tempfile.gettempdir()) / "main.xlsx"
        ExcelAggregator.create_main_workbook(csv_files, output, "main")
        print("[DRIVE]", self.drive.upload_or_update_file(output, "main.xlsx", period_folder_id, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        cleanup_paths(csv_files + [output])

    def _rebuild_global_main(self, root_id: str) -> None:
        all_csvs: list[Path] = []
        for folder in self.drive.list_folders(root_id):
            if re.fullmatch(r"\d+-\d+", folder["name"]):
                all_csvs.extend(download_csv_files_from_folder(self.drive, folder["id"]))

        output = Path(tempfile.gettempdir()) / "main_total.xlsx"
        ExcelAggregator.create_main_workbook(all_csvs, output, "main_total")
        print("[DRIVE]", self.drive.upload_or_update_file(output, "main_total.xlsx", root_id, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
        cleanup_paths(all_csvs + [output])


def cleanup_paths(paths: list[Path]) -> None:
    for p in paths:
        p.unlink(missing_ok=True)


def download_csv_files_from_folder(drive: DriveClient, folder_id: str) -> list[Path]:
    files = drive.list_files(folder_id, mime_exclude_folder=True)
    csv_candidates = [f for f in files if f["name"].lower().endswith(".csv")]
    local_paths: list[Path] = []
    for item in csv_candidates:
        target = Path(tempfile.gettempdir()) / f"{item['id']}_{item['name']}"
        drive.download_file(item["id"], target)
        local_paths.append(target)
    return local_paths


def day_to_period_label(day_value: int) -> str:
    start = ((day_value - 1) // 60) * 60 + 1
    end = start + 59
    return f"{start}-{end}"


def extract_day_from_csv(csv_file: Path) -> Optional[int]:
    with csv_file.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        first_row = next(reader, None)
        if not first_row:
            return None

        if len(first_row) > 1 and first_row[1].strip().isdigit():
            return int(first_row[1].strip())

        data_row = next(reader, None)
        if data_row and len(data_row) > 1 and data_row[1].strip().isdigit():
            return int(data_row[1].strip())
    return None


def is_game_running(process_names: tuple[str, ...]) -> bool:
    expected = {name.lower() for name in process_names}
    for proc in psutil.process_iter(attrs=["name"]):
        name = (proc.info.get("name") or "").lower()
        if name in expected:
            return True
    return False


def find_latest_save_folder(savegames_root: Path) -> Optional[Path]:
    if not savegames_root.exists():
        return None
    version_dirs = [p for p in savegames_root.iterdir() if p.is_dir()]
    if not version_dirs:
        return None
    latest_version = max(version_dirs, key=lambda p: p.stat().st_mtime)
    save_dirs = [p for p in latest_version.iterdir() if p.is_dir()]
    if not save_dirs:
        return None
    return max(save_dirs, key=lambda p: p.stat().st_mtime)


def build_default_config() -> Config:
    user_profile = os.environ.get("USERPROFILE")
    user_home = Path(user_profile) if user_profile else Path(os.path.expanduser("~"))
    savegames_root = (
        user_home
        / "AppData"
        / "LocalLow"
        / "Hovgaard Games"
        / "Big Ambitions"
        / "SaveGames"
    )
    script_dir = Path(__file__).resolve().parent
    service_account_file = Path(
        os.getenv("SERVICE_ACCOUNT_FILE", str(script_dir / "service_account_credentials.json"))
    )
    names = os.getenv("GAME_PROCESS_NAMES", "BigAmbitions.exe,UnityPlayer.exe")
    process_names = tuple(n.strip() for n in names.split(",") if n.strip())

    return Config(
        savegames_root=savegames_root,
        service_account_file=service_account_file,
        process_names=process_names,
    )


def preflight(config: Config) -> list[str]:
    errors: list[str] = []
    if os.name != "nt":
        errors.append("Bu script Windows path varsayımıyla yazıldı (os.name != 'nt').")
    if not config.savegames_root.exists():
        errors.append(f"SaveGames klasörü bulunamadı: {config.savegames_root}")
    if not config.service_account_file.exists():
        errors.append(f"Service account dosyası yok: {config.service_account_file}")
    if not config.process_names:
        errors.append("GAME_PROCESS_NAMES boş olamaz.")
    return errors


def run_loop(config: Config) -> None:
    drive = DriveClient(config.service_account_file)
    observer: Optional[Observer] = None
    watched_folder: Optional[Path] = None

    print("[INFO] Otomasyon başladı. Oyun süreci izleniyor...")
    while True:
        try:
            if is_game_running(config.process_names):
                latest_folder = find_latest_save_folder(config.savegames_root)
                if latest_folder is None:
                    print("[WARN] En güncel save klasörü bulunamadı.")
                else:
                    tx = latest_folder / "transactions.csv"
                    if tx.exists() and observer is None:
                        handler = TransactionsHandler(drive, config.file_settle_seconds)
                        observer = Observer()
                        observer.schedule(handler, str(latest_folder), recursive=False)
                        observer.start()
                        watched_folder = latest_folder
                        print(f"[INFO] İzleme başladı: {latest_folder}")
                    elif tx.exists() and watched_folder != latest_folder and observer is not None:
                        observer.stop()
                        observer.join(timeout=5)
                        handler = TransactionsHandler(drive, config.file_settle_seconds)
                        observer = Observer()
                        observer.schedule(handler, str(latest_folder), recursive=False)
                        observer.start()
                        watched_folder = latest_folder
                        print(f"[INFO] İzlenen klasör güncellendi: {latest_folder}")
            else:
                if observer is not None:
                    observer.stop()
                    observer.join(timeout=5)
                    observer = None
                    watched_folder = None
                    print("[INFO] Oyun kapalı, izleme durdu.")

            time.sleep(config.poll_seconds)
        except KeyboardInterrupt:
            print("[INFO] Çıkış sinyali alındı.")
            break
        except Exception as exc:
            print(f"[ERROR] Ana döngü hatası: {exc}")
            time.sleep(config.poll_seconds)

    if observer is not None:
        observer.stop()
        observer.join(timeout=5)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Big Ambitions -> Google Drive otomasyonu")
    parser.add_argument("--doctor", action="store_true", help="Sadece doğrulama yap ve çık")
    parser.add_argument("--no-gui", action="store_true", help="GUI açmadan çalıştır")
    return parser.parse_args()


def launch_config_gui(default_config: Config) -> Optional[Config]:
    result: dict[str, Config | None] = {"config": None}
    root = tk.Tk()
    root.title("Big Ambitions Drive Sync - Ayarlar")
    root.resizable(False, False)

    savegames_var = tk.StringVar(value=str(default_config.savegames_root))
    service_account_var = tk.StringVar(value=str(default_config.service_account_file))
    process_var = tk.StringVar(value=",".join(default_config.process_names))
    poll_var = tk.StringVar(value=str(default_config.poll_seconds))
    settle_var = tk.StringVar(value=str(default_config.file_settle_seconds))

    def browse_savegames() -> None:
        selected = filedialog.askdirectory(title="SaveGames klasörünü seç")
        if selected:
            savegames_var.set(selected)

    def browse_service_account() -> None:
        selected = filedialog.askopenfilename(
            title="Service account JSON seç",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if selected:
            service_account_var.set(selected)

    def on_start() -> None:
        try:
            poll_seconds = int(poll_var.get().strip())
            settle_seconds = int(settle_var.get().strip())
            process_names = tuple(p.strip() for p in process_var.get().split(",") if p.strip())
            cfg = Config(
                savegames_root=Path(savegames_var.get().strip()),
                service_account_file=Path(service_account_var.get().strip()),
                process_names=process_names,
                poll_seconds=poll_seconds,
                file_settle_seconds=settle_seconds,
            )
            errors = preflight(cfg)
            if errors:
                messagebox.showerror("Ayar Hatası", "\n".join(errors))
                return
            result["config"] = cfg
            root.destroy()
        except ValueError:
            messagebox.showerror("Ayar Hatası", "Süre değerleri sayı olmalı.")

    row = 0
    tk.Label(root, text="SaveGames klasörü").grid(row=row, column=0, sticky="w", padx=8, pady=6)
    tk.Entry(root, width=60, textvariable=savegames_var).grid(row=row, column=1, padx=8, pady=6)
    tk.Button(root, text="Seç", command=browse_savegames).grid(row=row, column=2, padx=8, pady=6)

    row += 1
    tk.Label(root, text="Service Account JSON").grid(row=row, column=0, sticky="w", padx=8, pady=6)
    tk.Entry(root, width=60, textvariable=service_account_var).grid(row=row, column=1, padx=8, pady=6)
    tk.Button(root, text="Seç", command=browse_service_account).grid(row=row, column=2, padx=8, pady=6)

    row += 1
    tk.Label(root, text="Process adları (virgülle)").grid(row=row, column=0, sticky="w", padx=8, pady=6)
    tk.Entry(root, width=60, textvariable=process_var).grid(row=row, column=1, padx=8, pady=6)

    row += 1
    tk.Label(root, text="Döngü bekleme sn").grid(row=row, column=0, sticky="w", padx=8, pady=6)
    tk.Entry(root, width=20, textvariable=poll_var).grid(row=row, column=1, sticky="w", padx=8, pady=6)

    row += 1
    tk.Label(root, text="Dosya yazım bekleme sn").grid(row=row, column=0, sticky="w", padx=8, pady=6)
    tk.Entry(root, width=20, textvariable=settle_var).grid(row=row, column=1, sticky="w", padx=8, pady=6)

    row += 1
    tk.Button(root, text="Başlat", command=on_start, width=18).grid(row=row, column=1, sticky="w", padx=8, pady=12)
    tk.Button(root, text="İptal", command=root.destroy, width=12).grid(row=row, column=1, sticky="e", padx=8, pady=12)

    root.mainloop()
    built = result["config"]
    return built if isinstance(built, Config) else None


def main() -> None:
    args = parse_args()
    default_config = build_default_config()
    config = default_config if args.no_gui else launch_config_gui(default_config)
    if config is None:
        print("[INFO] İşlem iptal edildi.")
        return

    errors = preflight(config)
    if errors:
        print("[PRECHECK] Hazır değil:")
        for issue in errors:
            print(" -", issue)
        raise SystemExit(1)

    print("[PRECHECK] Hazır.")
    if args.doctor:
        return

    run_loop(config)


if __name__ == "__main__":
    main()
